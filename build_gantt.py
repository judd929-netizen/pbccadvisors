"""Build an editable deal-timeline Gantt chart for Excel.

Bars are drawn natively by Excel conditional formatting from the Start/End/%Complete
columns -- edit those cells and the bars redraw automatically. No macros, no chart object.
"""
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# ---- Brand palette (PBG/PBC blue/white/gray) ----
NAVY      = "1F3A5F"   # headers
BLUE      = "2E6DB4"   # planned bar
BLUE_DK   = "16406E"   # completed portion of bar
BLUE_LT   = "DCE6F2"   # phase band / light fill
GRAY_LT   = "F2F4F7"   # zebra
GRAY_MED  = "B8C0CC"   # grid lines
TODAY_CLR = "E8A33D"   # today marker (amber accent)
WHITE     = "FFFFFF"

thin = Side(style="thin", color=GRAY_MED)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def font(sz=10, bold=False, color="000000", name="Arial"):
    return Font(name=name, size=sz, bold=bold, color=color)

def fill(hexcol):
    return PatternFill("solid", fgColor=hexcol)

# ---- Timeline config ----
PROJECT_START = dt.date(2026, 8, 3)   # a Monday
N_WEEKS = 20
TODAY = dt.date(2026, 8, 2)

# ---- Task list: (phase, task, owner, start_week, end_week, pct_complete) ----
# Weeks are 1-indexed offsets from PROJECT_START; edit dates freely in the sheet later.
def wk(n):   # start date of week n (Monday)
    return PROJECT_START + dt.timedelta(weeks=n - 1)
def wk_end(n):  # Friday of week n
    return wk(n) + dt.timedelta(days=4)

PHASES = [
    ("Phase 1 - Engagement & Preparation", [
        ("Sign engagement letter",            "PBG / Client",  1, 1, 1.0),
        ("Financial data collection & QoE",   "Client / PBG",  1, 3, 0.6),
        ("Build financial model",             "PBG",           2, 4, 0.4),
        ("Draft CIM & marketing materials",   "PBG",           3, 5, 0.15),
    ]),
    ("Phase 2 - Marketing & Outreach", [
        ("Finalize buyer / lender list",      "PBG / Client",  4, 5, 0.0),
        ("Distribute teaser & NDAs",          "PBG",           5, 7, 0.0),
        ("Distribute CIM to signed parties",  "PBG",           6, 8, 0.0),
        ("Management presentations",          "PBG / Client",  7, 9, 0.0),
    ]),
    ("Phase 3 - Bids & Negotiation", [
        ("Receive IOIs / term sheets",        "Buyers",        9, 10, 0.0),
        ("Evaluate & shortlist",              "PBG / Client",  10, 11, 0.0),
        ("Negotiate & sign LOI",              "PBG / Legal",   11, 13, 0.0),
    ]),
    ("Phase 4 - Diligence & Documentation", [
        ("Buyer due diligence",               "Buyer / Client",13, 17, 0.0),
        ("Draft definitive agreements",       "Legal",         14, 17, 0.0),
        ("Negotiate purchase agreement",      "PBG / Legal",   15, 18, 0.0),
    ]),
    ("Phase 5 - Closing", [
        ("Final approvals & funding",         "All parties",   17, 19, 0.0),
        ("Sign & close",                      "All parties",   19, 20, 0.0),
    ]),
]

wb = Workbook()
ws = wb.active
ws.title = "Deal Timeline"
ws.sheet_view.showGridLines = False

# Column layout
COL_TASK, COL_OWNER, COL_START, COL_END, COL_DAYS, COL_PCT = 1, 2, 3, 4, 5, 6
FIRST_GRID_COL = 7  # column G

# ---- Title block ----
ws.cell(1, 1, "Deal Timeline — Gantt Tracker").font = font(16, True, NAVY)
ws.cell(2, 1, "PBG Capital Advisors  ·  edit Start / End / % Complete and the bars redraw automatically").font = font(9, False, "5A6472")

# Legend (row 3)
legend = [
    ("■ Planned",   BLUE),
    ("■ Completed", BLUE_DK),
    ("▮ This week", TODAY_CLR),
]
lc = 1
for text, col in legend:
    c = ws.cell(3, lc, text)
    c.font = font(9, True, col)
    lc += 2

HEADER_ROW = 5
DATA_START = HEADER_ROW + 1

# ---- Header row ----
headers = [(COL_TASK, "Task"), (COL_OWNER, "Owner"), (COL_START, "Start"),
           (COL_END, "End"), (COL_DAYS, "Days"), (COL_PCT, "% Done")]
for col, label in headers:
    c = ws.cell(HEADER_ROW, col, label)
    c.font = font(10, True, WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_all

# Week grid headers (dates)
for i in range(N_WEEKS):
    col = FIRST_GRID_COL + i
    d = wk(i + 1)
    c = ws.cell(HEADER_ROW, col, d)
    c.number_format = "mmm d"
    c.font = font(8, True, WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
    c.border = border_all

LAST_GRID_COL = FIRST_GRID_COL + N_WEEKS - 1
last_letter = get_column_letter(LAST_GRID_COL)
first_letter = get_column_letter(FIRST_GRID_COL)

# ---- Write phases & tasks ----
r = DATA_START
task_rows = []
for phase_name, tasks in PHASES:
    # Phase band row
    ws.cell(r, COL_TASK, phase_name).font = font(10, True, NAVY)
    for col in range(1, LAST_GRID_COL + 1):
        cell = ws.cell(r, col)
        cell.fill = fill(BLUE_LT)
        cell.border = border_all
    r += 1
    for task, owner, sw, ew, pct in tasks:
        ws.cell(r, COL_TASK, "   " + task).font = font(10)
        ws.cell(r, COL_OWNER, owner).font = font(9, color="5A6472")
        cs = ws.cell(r, COL_START, wk(sw));  cs.number_format = "mmm d"; cs.font = font(9)
        ce = ws.cell(r, COL_END, wk_end(ew)); ce.number_format = "mmm d"; ce.font = font(9)
        cd = ws.cell(r, COL_DAYS, f"=NETWORKDAYS({get_column_letter(COL_START)}{r},{get_column_letter(COL_END)}{r})")
        cd.font = font(9); cd.alignment = Alignment(horizontal="center")
        cp = ws.cell(r, COL_PCT, pct); cp.number_format = "0%"; cp.font = font(9)
        cp.alignment = Alignment(horizontal="center")
        for col in range(1, LAST_GRID_COL + 1):
            ws.cell(r, col).border = border_all
        task_rows.append(r)
        r += 1

LAST_ROW = r - 1

# ---- Conditional formatting: draw the bars ----
grid_range = f"{first_letter}{DATA_START}:{last_letter}{LAST_ROW}"
sc, ec, pc = (get_column_letter(COL_START), get_column_letter(COL_END), get_column_letter(COL_PCT))

# Completed portion (higher priority): week overlaps task AND week-start is within the % done span.
completed = (f"AND(${sc}{DATA_START}<>\"\","
             f"{first_letter}${HEADER_ROW}+4>=${sc}{DATA_START},"
             f"{first_letter}${HEADER_ROW}<=${ec}{DATA_START},"
             f"{first_letter}${HEADER_ROW}<=${sc}{DATA_START}+(${ec}{DATA_START}-${sc}{DATA_START})*${pc}{DATA_START})")
ws.conditional_formatting.add(grid_range,
    FormulaRule(formula=[completed], fill=fill(BLUE_DK), stopIfTrue=True))

# Planned bar: any week that overlaps [Start, End].
planned = (f"AND(${sc}{DATA_START}<>\"\","
           f"{first_letter}${HEADER_ROW}+4>=${sc}{DATA_START},"
           f"{first_letter}${HEADER_ROW}<=${ec}{DATA_START})")
ws.conditional_formatting.add(grid_range,
    FormulaRule(formula=[planned], fill=fill(BLUE)))

# Today marker: highlight the header cell of the week containing TODAY().
hdr_range = f"{first_letter}{HEADER_ROW}:{last_letter}{HEADER_ROW}"
today_rule = f"AND(TODAY()>={first_letter}{HEADER_ROW},TODAY()<{first_letter}{HEADER_ROW}+7)"
ws.conditional_formatting.add(hdr_range,
    FormulaRule(formula=[today_rule], fill=fill(TODAY_CLR), stopIfTrue=True))
# Faint today column down the grid
today_col = f"AND(TODAY()>={first_letter}${HEADER_ROW},TODAY()<{first_letter}${HEADER_ROW}+7)"
ws.conditional_formatting.add(grid_range,
    FormulaRule(formula=[today_col], fill=fill("FBEBD4")))

# ---- Dimensions ----
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 9
ws.column_dimensions["D"].width = 9
ws.column_dimensions["E"].width = 6
ws.column_dimensions["F"].width = 7
for i in range(N_WEEKS):
    ws.column_dimensions[get_column_letter(FIRST_GRID_COL + i)].width = 3.6
ws.row_dimensions[HEADER_ROW].height = 46

# Freeze panes: keep task labels + header visible while scrolling the grid
ws.freeze_panes = ws.cell(DATA_START, FIRST_GRID_COL).coordinate

# Data bar on % column for a quick progress read
from openpyxl.formatting.rule import DataBarRule
ws.conditional_formatting.add(
    f"{get_column_letter(COL_PCT)}{DATA_START}:{get_column_letter(COL_PCT)}{LAST_ROW}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                color=BLUE, showValue=True))

# ---- How-to note below the table ----
note_row = LAST_ROW + 2
ws.cell(note_row, 1, "How to use").font = font(10, True, NAVY)
notes = [
    "• Edit the Start and End dates (cols C–D) — the blue bars redraw automatically.",
    "• Set % Done (col F) — the darker segment fills the completed portion of each bar.",
    "• The amber column marks the week containing today's date (updates live via TODAY()).",
    "• Add a task: insert a row, type the name/owner/dates; bars appear with no extra setup.",
    "• Extend the timeline: the grid runs " + wk(1).strftime("%b %d") + " to " + wk(N_WEEKS).strftime("%b %d") + " (20 weeks).",
]
for i, n in enumerate(notes):
    ws.cell(note_row + 1 + i, 1, n).font = font(9, color="5A6472")

wb.save("Deal_Timeline_Gantt.xlsx")
print("saved Deal_Timeline_Gantt.xlsx  rows:", DATA_START, "-", LAST_ROW)
